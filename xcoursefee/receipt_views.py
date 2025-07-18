from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db.models import Sum
from django.utils import timezone
from decimal import Decimal
import io
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

from .models import StudentEnrollment, Payment, KitFee


@login_required
def enrollment_receipt_print(request, enrollment_id):
    """Print-friendly receipt for enrollment with all fees"""
    enrollment = get_object_or_404(StudentEnrollment, pk=enrollment_id)
    payments = enrollment.payments.filter(status='completed').order_by('payment_date')
    kit_fees = KitFee.objects.filter(
        enrollment=enrollment, 
        payment_status='paid'
    ).select_related('course_kit__kit')
    
    # Calculate totals
    total_course_fees = enrollment.get_total_fees()
    total_kit_fees = enrollment.get_total_kit_fees()
    total_paid = enrollment.get_total_paid()
    total_kit_paid = kit_fees.aggregate(total=Sum('amount'))['total'] or Decimal('0.000')
    grand_total = total_course_fees + total_kit_fees
    total_paid_all = total_paid + total_kit_paid
    
    context = {
        'enrollment': enrollment,
        'payments': payments,
        'kit_fees': kit_fees,
        'total_course_fees': total_course_fees,
        'total_kit_fees': total_kit_fees,
        'total_paid': total_paid,
        'total_kit_paid': total_kit_paid,
        'grand_total': grand_total,
        'total_paid_all': total_paid_all,
        'outstanding_balance': grand_total - total_paid_all,
        'receipt_date': timezone.now().date(),
    }
    return render(request, 'xcoursefee/receipt_print.html', context)


@login_required
def enrollment_receipt_pdf(request, enrollment_id):
    """Generate PDF receipt for enrollment"""
    enrollment = get_object_or_404(StudentEnrollment, pk=enrollment_id)
    payments = enrollment.payments.filter(status='completed').order_by('payment_date')
    kit_fees = KitFee.objects.filter(
        enrollment=enrollment, 
        payment_status='paid'
    ).select_related('course_kit__kit')
    
    # Create PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="receipt_{enrollment.student.student_name.replace(" ", "_")}_{enrollment.course.course_code}.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []
    
    # Header
    header_style = ParagraphStyle(
        'HeaderStyle',
        parent=styles['Heading1'],
        fontSize=20,
        textColor=colors.darkblue,
        alignment=1,  # Center
        spaceAfter=30,
    )
    story.append(Paragraph("EduPulse - Fee Receipt", header_style))
    
    # Receipt info
    info_style = styles['Normal']
    story.append(Paragraph(f"<b>Receipt Date:</b> {timezone.now().date().strftime('%B %d, %Y')}", info_style))
    story.append(Paragraph(f"<b>Student:</b> {enrollment.student.student_name}", info_style))
    story.append(Paragraph(f"<b>Student ID:</b> {enrollment.student.student_id}", info_style))
    story.append(Paragraph(f"<b>Course:</b> {enrollment.course.name} ({enrollment.course.course_code})", info_style))
    story.append(Paragraph(f"<b>Enrollment Date:</b> {enrollment.enrollment_date.strftime('%B %d, %Y')}", info_style))
    story.append(Spacer(1, 20))
    
    # Course Fees Table
    if payments.exists():
        story.append(Paragraph("<b>Course Fee Payments</b>", styles['Heading3']))
        payment_data = [['Date', 'Amount (KWD)', 'Method', 'Receipt #']]
        for payment in payments:
            payment_data.append([
                payment.payment_date.strftime('%m/%d/%Y'),
                str(payment.amount),
                payment.get_payment_method_display(),
                payment.receipt_number or '-'
            ])
        
        payment_table = Table(payment_data)
        payment_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(payment_table)
        story.append(Spacer(1, 20))
    
    # Kit Fees Table
    if kit_fees.exists():
        story.append(Paragraph("<b>Kit Fee Payments</b>", styles['Heading3']))
        kit_data = [['Kit Name', 'Amount (KWD)', 'Payment Date', 'Delivery Status']]
        for kit_fee in kit_fees:
            kit_data.append([
                kit_fee.course_kit.kit.name,
                str(kit_fee.amount),
                kit_fee.payment_date.strftime('%m/%d/%Y') if kit_fee.payment_date else '-',
                kit_fee.get_delivery_status_display()
            ])
        
        kit_table = Table(kit_data)
        kit_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgreen),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(kit_table)
        story.append(Spacer(1, 20))
    
    # Summary
    total_course_fees = enrollment.get_total_fees()
    total_kit_fees = enrollment.get_total_kit_fees()
    total_paid = enrollment.get_total_paid()
    total_kit_paid = kit_fees.aggregate(total=Sum('amount'))['total'] or Decimal('0.000')
    grand_total = total_course_fees + total_kit_fees
    total_paid_all = total_paid + total_kit_paid
    
    story.append(Paragraph("<b>Payment Summary</b>", styles['Heading3']))
    summary_data = [
        ['Total Course Fees:', f'{total_course_fees} KWD'],
        ['Total Kit Fees:', f'{total_kit_fees} KWD'],
        ['Grand Total:', f'{grand_total} KWD'],
        ['Total Paid:', f'{total_paid_all} KWD'],
        ['Outstanding Balance:', f'{grand_total - total_paid_all} KWD'],
    ]
    
    summary_table = Table(summary_data)
    summary_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 12),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    story.append(summary_table)
    
    # Footer
    story.append(Spacer(1, 30))
    footer_style = ParagraphStyle(
        'FooterStyle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.grey,
        alignment=1,  # Center
    )
    story.append(Paragraph("Thank you for choosing EduPulse!", footer_style))
    story.append(Paragraph("This is an official receipt for your records.", footer_style))
    
    doc.build(story)
    return response


@login_required
def enrollment_receipt_excel(request, enrollment_id):
    """Generate Excel receipt for enrollment"""
    enrollment = get_object_or_404(StudentEnrollment, pk=enrollment_id)
    payments = enrollment.payments.filter(status='completed').order_by('payment_date')
    kit_fees = KitFee.objects.filter(
        enrollment=enrollment, 
        payment_status='paid'
    ).select_related('course_kit__kit')
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Fee Receipt"
    
    # Styles
    header_font = Font(size=16, bold=True, color="1F4E79")
    label_font = Font(bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Header
    ws['A1'] = "EduPulse - Fee Receipt"
    ws['A1'].font = header_font
    ws.merge_cells('A1:D1')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Student Information
    row = 3
    ws[f'A{row}'] = "Student Information"
    ws[f'A{row}'].font = label_font
    row += 1
    
    info_data = [
        ("Receipt Date:", timezone.now().date().strftime('%B %d, %Y')),
        ("Student Name:", enrollment.student.student_name),
        ("Student ID:", enrollment.student.student_id),
        ("Course:", f"{enrollment.course.name} ({enrollment.course.course_code})"),
        ("Enrollment Date:", enrollment.enrollment_date.strftime('%B %d, %Y')),
    ]
    
    for label, value in info_data:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = label_font
        ws[f'B{row}'] = value
        row += 1
    
    row += 1
    
    # Course Payments
    if payments.exists():
        ws[f'A{row}'] = "Course Fee Payments"
        ws[f'A{row}'].font = label_font
        row += 1
        
        headers = ['Date', 'Amount (KWD)', 'Payment Method', 'Receipt Number']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = label_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        row += 1
        
        for payment in payments:
            data = [
                payment.payment_date.strftime('%m/%d/%Y'),
                float(payment.amount),
                payment.get_payment_method_display(),
                payment.receipt_number or '-'
            ]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                if col == 2:  # Amount column
                    cell.alignment = Alignment(horizontal='right')
            row += 1
        row += 1
    
    # Kit Payments
    if kit_fees.exists():
        ws[f'A{row}'] = "Kit Fee Payments"
        ws[f'A{row}'].font = label_font
        row += 1
        
        headers = ['Kit Name', 'Amount (KWD)', 'Payment Date', 'Delivery Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = label_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        row += 1
        
        for kit_fee in kit_fees:
            data = [
                kit_fee.course_kit.kit.name,
                float(kit_fee.amount),
                kit_fee.payment_date.strftime('%m/%d/%Y') if kit_fee.payment_date else '-',
                kit_fee.get_delivery_status_display()
            ]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                if col == 2:  # Amount column
                    cell.alignment = Alignment(horizontal='right')
            row += 1
        row += 1
    
    # Summary
    ws[f'A{row}'] = "Payment Summary"
    ws[f'A{row}'].font = label_font
    row += 1
    
    total_course_fees = enrollment.get_total_fees()
    total_kit_fees = enrollment.get_total_kit_fees()
    total_paid = enrollment.get_total_paid()
    total_kit_paid = kit_fees.aggregate(total=Sum('amount'))['total'] or Decimal('0.000')
    grand_total = total_course_fees + total_kit_fees
    total_paid_all = total_paid + total_kit_paid
    
    summary_data = [
        ("Total Course Fees:", float(total_course_fees)),
        ("Total Kit Fees:", float(total_kit_fees)),
        ("Grand Total:", float(grand_total)),
        ("Total Paid:", float(total_paid_all)),
        ("Outstanding Balance:", float(grand_total - total_paid_all)),
    ]
    
    for label, value in summary_data:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = label_font
        ws[f'B{row}'] = f"{value} KWD"
        ws[f'B{row}'].alignment = Alignment(horizontal='right')
        row += 1
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="receipt_{enrollment.student.student_name.replace(" ", "_")}_{enrollment.course.course_code}.xlsx"'
    
    wb.save(response)
    return response


@login_required
def payment_receipt_print(request, payment_id):
    """Print-friendly receipt for individual payment"""
    payment = get_object_or_404(Payment, pk=payment_id)
    
    context = {
        'payment': payment,
        'receipt_date': timezone.now().date(),
    }
    return render(request, 'xcoursefee/payment_receipt_print.html', context)


@login_required
def payment_receipt_pdf(request, payment_id):
    """Generate PDF receipt for individual payment"""
    payment = get_object_or_404(Payment, pk=payment_id)
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="payment_receipt_{payment.id}_{payment.payment_date.strftime("%Y%m%d")}.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []
    
    # Header
    header_style = ParagraphStyle(
        'HeaderStyle',
        parent=styles['Heading1'],
        fontSize=20,
        textColor=colors.darkblue,
        alignment=1,  # Center
        spaceAfter=30,
    )
    story.append(Paragraph("EduPulse - Payment Receipt", header_style))
    
    # Receipt info
    info_style = styles['Normal']
    story.append(Paragraph(f"<b>Receipt Date:</b> {timezone.now().date().strftime('%B %d, %Y')}", info_style))
    story.append(Paragraph(f"<b>Payment Date:</b> {payment.payment_date.strftime('%B %d, %Y')}", info_style))
    story.append(Paragraph(f"<b>Student:</b> {payment.enrollment.student.student_name}", info_style))
    story.append(Paragraph(f"<b>Course:</b> {payment.enrollment.course.name} ({payment.enrollment.course.course_code})", info_style))
    story.append(Paragraph(f"<b>Amount:</b> {payment.amount} KWD", info_style))
    story.append(Paragraph(f"<b>Payment Method:</b> {payment.get_payment_method_display()}", info_style))
    if payment.reference_number:
        story.append(Paragraph(f"<b>Reference Number:</b> {payment.reference_number}", info_style))
    if payment.receipt_number:
        story.append(Paragraph(f"<b>Receipt Number:</b> {payment.receipt_number}", info_style))
    story.append(Spacer(1, 30))
    
    # Footer
    footer_style = ParagraphStyle(
        'FooterStyle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.grey,
        alignment=1,  # Center
    )
    story.append(Paragraph("Thank you for your payment!", footer_style))
    story.append(Paragraph("This is an official receipt for your records.", footer_style))
    
    doc.build(story)
    return response 